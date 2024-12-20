import os
import base64
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk  # Importando ttk para a barra de progresso
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from time import sleep
from datetime import datetime
from anticaptchaofficial.recaptchav2proxyless import *
from chave_API import chave_api

def configurar_chrome_options():
    chrome_options = Options()
    chrome_options.add_argument('--kiosk-printing')  # Ativa a impressão automática sem diálogo
    return chrome_options

def gerar_pdf_dinamico(driver, caminho_diretorio, nome_pdf):
    # Define o caminho completo do arquivo PDF usando o nome_pdf fornecido
    caminho_arquivo_pdf = os.path.join(caminho_diretorio, f"{nome_pdf}.pdf")

    # Usa a API DevTools para gerar o PDF da página atual
    result = driver.execute_cdp_cmd("Page.printToPDF", {
        "landscape": False,                 # Define orientação do PDF como retrato
        "displayHeaderFooter": False,       # Não exibe cabeçalho e rodapé
        "printBackground": True,            # Inclui o fundo da página no PDF
        "preferCSSPageSize": True           # Usa o tamanho da página definido por CSS
    })

    # Decodifica o PDF gerado (base64) e salva em um arquivo
    with open(caminho_arquivo_pdf, "wb") as file:
        file.write(base64.b64decode(result['data']))

    print(f"PDF gerado e salvo em: {caminho_arquivo_pdf}")

def start_process(excel_path, sheet_name, start_row, output_dir, progress_var, progress_bar, log_text, demanda_nome):  # Adição de variáveis para progresso e log
    # Início da contagem de tempo
    start_time = datetime.now()
    
    # Carregar a planilha Excel
    workbook = openpyxl.load_workbook(excel_path)
    planilha = workbook.worksheets[0]

    column_index = 3
    total_rows = planilha.max_row
    row = start_row  # Use a variável start_row passada como argumento

    progress_step = 100 / (total_rows - start_row + 1)  # Definindo passo para a barra de progresso
    
    try:
        # Tenta salvar a planilha
        workbook.save(excel_path)
    except Exception as e:
        messagebox.showerror("Feche a Planilha", f"A planilha está aberta. Feche e clique no botão iniciar novamente")
        return
    
    # Renomear o arquivo Excel com o nome da demanda
    novo_nome_excel = os.path.join(os.path.dirname(excel_path), f"{demanda_nome}.xlsx")
    shutil.move(excel_path, novo_nome_excel)
    excel_path = novo_nome_excel  # Atualizar o caminho do arquivo Excel com o novo nome

    log_text.insert(tk.END, f"Arquivo Excel renomeado para: {novo_nome_excel}\n")
    log_text.see(tk.END)

    # Configurando o Chrome com as opções de impressão
    chrome_options = configurar_chrome_options()

    # Iniciar o Chrome sem precisar especificar o chromedriver
    driver = webdriver.Chrome(options=chrome_options)

    # Acessar o link desejado
    link = 'https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf'
    driver.get(link)
    sleep(2)

    while row <= total_rows:
        # Atualizando a barra de progresso e mostrando a linha atual no console
        progress_var.set(progress_var.get() + progress_step)
        progress_bar.update()

        log_text.insert(tk.END, f"Processando linha {row}/{total_rows}\n")  # Log na interface para mostrar a linha processada
        log_text.see(tk.END)  # Faz o scroll do log automaticamente para a última linha

        # Verificação se existe um "OK" na coluna 21 da linha atual
        validador = planilha.cell(row=row, column=21).value
            
        if validador == "OK":
            log_text.insert(tk.END, f"Linha {row}: Já processada (OK).\n")
            row += 1
            continue
        
        Num_CDA = planilha.cell(row=row, column=column_index).value

        if Num_CDA is None:
            log_text.insert(tk.END, "Num_CDA está vazio. Saindo do loop.\n")
            break
        
        log_text.insert(tk.END, f"Linha {row}: Consultando CDA {Num_CDA}...\n")

        Campo_CDA = driver.find_element(By.ID, "consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta")
        Campo_CDA.clear()
        Campo_CDA.send_keys(Num_CDA)

        btn_Consultar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]")
        btn_Consultar.click()
        sleep(1)
        
        try:
            captchanaovalidado = driver.find_element(By.XPATH, "//*[@id='messages']/tbody/tr/td/span[2]")
            if "Recaptcha não validado" in captchanaovalidado.text:
                chave_captcha = driver.find_element(By.ID, 'recaptcha').get_attribute('data-sitekey')

                solver = recaptchaV2Proxyless()
                solver.set_verbose(1)
                solver.set_key(chave_api)
                solver.set_website_url(link)
                solver.set_website_key(chave_captcha)

                resposta = solver.solve_and_return_solution()
                
                if resposta != 0:
                    print(resposta)
                    # preencher o campo do token do captcha
                    # g-recaptcha-response
                    driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")
                    #driver.find_element(By.ID, 'recaptcha-verify-button').click()
                    primeiro_registro = False
                    btn_Consultar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]")
                    btn_Consultar.click()
                    sleep(1)
                else:
                    print(solver.err_string)

        except:
            sleep(0)

        resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
        resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

        if "Nenhum resultado com os critérios de consulta" in resultado_msg:
            log_text.insert(tk.END, f"Linha {row}: Nenhum resultado com os critérios de consulta.\n")
            planilha.cell(row=row, column=column_index + 1, value="Nenhum resultado com os critérios de consulta")
            planilha.cell(row=row, column=21, value="OK")  # Escreve "OK" na linha atual na coluna 21 (Coluna de verificação)
            workbook.save(excel_path)
            
            # Aqui você chama a função para gerar e salvar o PDF com o nome da CDA
            gerar_pdf_dinamico(driver, output_dir, Num_CDA)
            
            row += 1
            continue

        log_text.insert(tk.END, f"Linha {row}: Consultando mais informações...\n")

        # Consultar IPVA
        href_IPVA = driver.find_element(By.ID, "consultaDebitoForm:dataTable:0:lnkConsultaDebito")
        href_IPVA.click()
        sleep(1)

        # Consultar Registro
        href_Registro = driver.find_element(By.XPATH, "//a[@href='#']")
        href_Registro.click()
        sleep(1)  
        
        try:
            valor_honorarios_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1134:4:j_id1142")
            valor_honorarios = valor_honorarios_element.text
        except:
            valor_honorarios = 0

        try:
            valor_mora_multa_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1134:5:j_id1142")
            valor_mora_multa = valor_mora_multa_element.text
        except:
            valor_mora_multa = 0

        num_registro_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1028")
        num_registro = num_registro_element.find_element(By.TAG_NAME, "span").text
        numero_processo_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1044")
        numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
        numero_processo_outros_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1053")
        numero_processo_outros = numero_processo_outros_element.find_element(By.TAG_NAME, "span").text
        data_inscricao_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1036")
        data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text
        situacao_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1110")
        situacao = situacao_element.find_element(By.TAG_NAME, "span").text
        saldo_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1126")
        saldo = saldo_element.find_element(By.TAG_NAME, "span").text
        valor_principal_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1134:0:j_id1142")
        valor_principal = valor_principal_element.text
        valor_juros_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1134:2:j_id1142")
        valor_juros = valor_juros_element.text
        valor_multa_element = driver.find_element(By.ID, "consultaDebitoForm:j_id1134:3:j_id1142")
        valor_multa = valor_multa_element.text
        placa_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1193")
        placa = placa_element.text
        renavam_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1195")
        renavam = renavam_element.text
        chassi_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1197")
        chassi = chassi_element.text
        marca_modelo_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1199")
        marca_modelo = marca_modelo_element.text
        ano_fab_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1201")
        ano_fab = ano_fab_element.text
        ano_exercicio_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1203")
        ano_exercicio = ano_exercicio_element.text
        dt_parcelas_element = driver.find_element(By.ID, "consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1205")
        dt_parcelas = dt_parcelas_element.text

        log_text.insert(tk.END, f"Linha {row}: Dados obtidos - Placa: {placa}, Renavam: {renavam}, Chassi: {chassi}\n")
        
        coluna_atual = column_index + 1
        planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=numero_processo)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=numero_processo_outros)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=situacao)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=saldo)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=valor_principal)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=valor_juros)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=valor_multa)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=valor_mora_multa)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=placa)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=renavam)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=chassi)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=marca_modelo)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=ano_fab)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=ano_exercicio)
        coluna_atual += 1
        planilha.cell(row=row, column=coluna_atual, value=dt_parcelas)
        planilha.cell(row=row, column=21, value="OK")  # Escreve "OK" na linha atual na coluna 21 (Coluna de verificação)
        workbook.save(excel_path)
        CDA = str(Num_CDA)
        
        # Gera e salva o PDF com o nome da CDA
        gerar_pdf_dinamico(driver, output_dir, Num_CDA)

        # Botão Voltar
        wait = WebDriverWait(driver, 15)
        btn_Voltar = wait.until(EC.presence_of_element_located((By.ID, "consultaDebitoForm:btnVoltarDetalheDebito")))
        btn_Voltar.click()

        # Botão Voltar
        wait = WebDriverWait(driver, 15)
        btn_Voltar1 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:consultaDebito']/div[2]/input")))
        btn_Voltar1.click()
        row += 1
    
    # Criar uma pasta com o nome da demanda no diretório de saída
    demanda_dir = os.path.join(output_dir, demanda_nome)
    os.makedirs(demanda_dir, exist_ok=True)
    
    # Mover PDFs gerados e o Excel renomeado para a pasta criada
    for file in os.listdir(output_dir):
        if file.endswith(".pdf") or file == f"{demanda_nome}.xlsx":
            shutil.move(os.path.join(output_dir, file), os.path.join(demanda_dir, file))
    
    log_text.insert(tk.END, f"Arquivos movidos para a pasta: {demanda_dir}\n")
    log_text.see(tk.END)
    
    # Fim da contagem de tempo
    end_time = datetime.now()
    total_time = end_time - start_time

    total_seconds = total_time.total_seconds()
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    formatted_time = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

    driver.quit()
    messagebox.showinfo("Processo concluído", f"Todas as CDAs estão preenchidas com sucesso!\n\nProcesso iniciado às: {start_time.strftime('%d-%m-%Y %H:%M:%S')}\nProcesso finalizado às: {end_time.strftime('%d-%m-%Y %H:%M:%S')}\n\nTempo total: {formatted_time}")

def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    entry.delete(0, tk.END)
    entry.insert(0, filename)

def browse_directory(entry):
    directory = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, directory)

def main():
    root = tk.Tk()
    root.title("Automação PGE SP")

    tk.Label(root, text="Arquivo Excel:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=5)
    excel_entry = tk.Entry(root, width=50)
    excel_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="Procurar", command=lambda: browse_file(excel_entry)).grid(row=0, column=2, padx=(10, 150), pady=5, sticky=tk.E)

    tk.Label(root, text="Nome da Demanda:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=5)
    demanda_entry = tk.Entry(root, width=50)
    demanda_entry.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(root, text="Diretório de Saída dos PDFs:").grid(row=3, column=0, sticky=tk.W, padx=10, pady=5)
    output_dir_entry = tk.Entry(root, width=50)
    output_dir_entry.grid(row=3, column=1, padx=10, pady=5)
    tk.Button(root, text="Procurar", command=lambda: browse_directory(output_dir_entry)).grid(row=3, column=2, padx=(10, 150), pady=5, sticky=tk.E)

    # Adicionando a barra de progresso
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, orient=tk.HORIZONTAL, length=400, mode='determinate')
    progress_bar.grid(row=4, column=1, padx=10, pady=10, sticky=tk.W)

    # Adicionando uma área de log
    log_text = tk.Text(root, height=10, width=80)
    log_text.grid(row=5, column=0, columnspan=3, padx=10, pady=5)

    def on_start():
        excel_path = excel_entry.get()
        demanda_nome = demanda_entry.get()
        sheet_name = "IPVA SP"
        start_row = 3  # Início na linha 3 por padrão
        output_dir = output_dir_entry.get()

        if not os.path.exists(excel_path):
            messagebox.showerror("Erro", "O arquivo Excel não existe.")
            return
        if not os.path.exists(output_dir):
            messagebox.showerror("Erro", "O diretório de saída não existe.")
            return
        if not demanda_nome:
            messagebox.showerror("Erro", "Insira o nome da demanda.")
            return

        start_process(excel_path, sheet_name, start_row, output_dir, progress_var, progress_bar, log_text, demanda_nome)  # Passando a barra de progresso, o log, e o nome da demanda

    tk.Button(root, text="Iniciar", command=on_start).grid(row=6, column=1, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
